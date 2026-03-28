"""
File conversion utilities for the Estimator module.

Ported from BidParser (archived/parser/bidtabs/ilutility.py) and modernized.

Provides:
  - IDOT Pay Item Awards Excel (.xlsx) → standardized CSV
  - ISTHA bid tab Excel (.xlsx) → standardized CSV
  - Bid tab file renaming (standardized naming from file content)
  - Batch conversion for directories and GCS buckets
"""
from __future__ import annotations

import csv
import io
import logging
import re
from datetime import datetime
from pathlib import Path

import openpyxl

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# IDOT Pay Item Awards: Excel → CSV
# ---------------------------------------------------------------------------

# Standardized CSV headers for IDOT awards
# Matches the format expected by the existing IDOT awards parser
IDOT_AWARDS_HEADERS = [
    "code", "description", "unit", "quantity", "county",
    "district", "contract", "item", "price",
]

# IDOT Pay Item Reports come in three Excel formats over the years:
#
# Format A (2018 – mid-2021): 9 columns
#   0: PayItem, 1: Description, 2: UOM, 3: Quantity, 4: County,
#   5: District, 6: Contract, 7: Item, 8: Price
#   → Already in standardized order
#
# Format B (mid-2021 – 2022): 11 columns
#   Same as Format A but with 2 trailing null columns
#   → Same mapping as A, ignore trailing nulls
#
# Format C (2023+): 10 columns
#   0: PayItem, 1: Description, 2: UOM, 3: Quantity, 4: District,
#   5: Contract, 6: Item#, 7: Unit Price, 8: Awarded (Y/blank), 9: County
#   → County and District swapped, "Awarded" column inserted
#
# All remap to: code, description, unit, quantity, county, district, contract, item, price

_COL_MAP_FORMAT_A = [0, 1, 2, 3, 4, 5, 6, 7, 8]   # 9 or 11 cols, county@4 district@5
_COL_MAP_FORMAT_C = [0, 1, 2, 3, 9, 4, 5, 6, 7]    # 10 cols, county@9 district@4


def idot_awards_xlsx_to_csv(
    xlsx_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    """
    Convert an IDOT Pay Item Report with Awarded Prices Excel file to
    a standardized CSV.

    Input:  IDOT_PayItemAwards_2025_11_07.xlsx  (or legacy APIR110725.xlsx)
    Output: AWD_IL_IDOT_2025_11_07.csv

    The output CSV has headers:
        code, description, unit, quantity, county, district, contract, item, price

    Rows starting with "pay", "number", or blank in column A are skipped
    (these are Excel header/footer rows).

    Args:
        xlsx_path: Path to the input .xlsx file.
        output_path: Directory for output CSV. Defaults to same directory as input.

    Returns:
        Path to the created CSV file.
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_path) if output_path else xlsx_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)

    # Parse letting date from filename
    letting_date = _parse_date_from_filename(xlsx_path.stem)
    if letting_date:
        csv_name = f"AWD_IL_IDOT_{letting_date.year}_{letting_date.month:02d}_{letting_date.day:02d}.csv"
    else:
        # Fallback: use the stem with .csv extension
        csv_name = xlsx_path.stem + ".csv"
        logger.warning(f"Could not parse date from {xlsx_path.name}, using fallback name: {csv_name}")

    csv_path = output_dir / csv_name

    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active

    # Detect format by reading header rows
    col_map = _detect_format(ws)

    rows_written = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        writer.writerow(IDOT_AWARDS_HEADERS)

        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if not row or len(row) == 0:
                continue
            # Skip header/empty rows
            first_cell = str(row[0]).strip().lower() if row[0] is not None else ""
            if not first_cell or first_cell.startswith(("pay", "number", "vp3")):
                continue

            # Pad row to max needed length
            max_idx = max(col_map) + 1
            padded = list(row) + [None] * (max_idx - len(row)) if len(row) < max_idx else list(row)
            mapped = [_clean_cell(padded[i]) for i in col_map]
            writer.writerow(mapped)
            rows_written += 1

    wb.close()
    logger.info(f"Converted {xlsx_path.name} → {csv_name} ({rows_written} rows)")
    return csv_path


def batch_convert_idot_awards(
    input_dir: str | Path,
    output_dir: str | Path | None = None,
) -> list[Path]:
    """
    Convert all IDOT awards .xlsx files in a directory to standardized CSVs.

    Args:
        input_dir: Directory containing .xlsx files.
        output_dir: Directory for output CSVs. Defaults to input_dir.

    Returns:
        List of created CSV file paths.
    """
    input_dir = Path(input_dir)
    output_dir = Path(output_dir) if output_dir else input_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    results: list[Path] = []
    xlsx_files = sorted(input_dir.glob("*.xlsx"))

    logger.info(f"Found {len(xlsx_files)} .xlsx files in {input_dir}")

    for xlsx_path in xlsx_files:
        try:
            csv_path = idot_awards_xlsx_to_csv(xlsx_path, output_dir)
            results.append(csv_path)
        except Exception as e:
            logger.error(f"Failed to convert {xlsx_path.name}: {e}")

    logger.info(f"Converted {len(results)} of {len(xlsx_files)} files")
    return results


# ---------------------------------------------------------------------------
# ISTHA Bid Tabs: Excel → CSV
# ---------------------------------------------------------------------------

ISTHA_HEADERS = [
    "contractor", "rank", "code", "abbreviation", "unit",
    "quantity", "price", "bid_total", "doc_total", "project_no",
]


def _parse_excel_date_cell(value: str) -> str:
    """Parse a date string like 'Open Date: 12/20/2012' → '20121220'."""
    parts = value.split(":")[1].strip().split("/")
    m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
    if y < 2000:
        y += 2000
    return f"{y}{m:02d}{d:02d}"


def _auth_total_lookup(ws, contractor_col: int) -> float:
    """Find the 'Auth' total for a contractor column in an ISTHA sheet."""
    for row_idx in range(5, 9):
        cell_val = ws.cell(row=row_idx, column=contractor_col).value
        if isinstance(cell_val, str) and cell_val.startswith("Auth"):
            total = ws.cell(row=row_idx, column=contractor_col + 1).value
            return float(total) if total else 0.0
    return 0.0


def istha_xlsx_to_csv(
    xlsx_path: str | Path,
    output_path: str | Path | None = None,
) -> Path:
    """
    Convert an ISTHA bid tab Excel file to a standardized CSV.

    The output CSV has headers:
        contractor, rank, code, abbreviation, unit, quantity, price,
        bid_total, doc_total, project_no

    Args:
        xlsx_path: Path to the input .xlsx file.
        output_path: Directory for output CSV. Defaults to same directory as input.

    Returns:
        Path to the created CSV file.
    """
    xlsx_path = Path(xlsx_path)
    output_dir = Path(output_path) if output_path else xlsx_path.parent

    wb = openpyxl.load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb.active

    # Extract metadata from fixed positions
    project_no = str(ws.cell(row=2, column=1).value or "").split(":")[1].strip()
    open_date = _parse_excel_date_cell(str(ws.cell(row=4, column=1).value or ""))
    begin_date = _parse_excel_date_cell(str(ws.cell(row=5, column=1).value or ""))

    # Find contractors (row 3, starting at column 5, every other column)
    contractors: list[str] = []
    bidder_totals: list[float] = []
    auth_totals: list[float] = []

    col = 5  # 1-indexed, column E
    while col < ws.max_column:
        name = ws.cell(row=3, column=col).value
        if not name:
            break
        contractors.append(str(name))
        bidder_totals.append(float(ws.cell(row=5, column=col + 1).value or 0))
        auth_totals.append(_auth_total_lookup(ws, col))
        col += 2

    # Build CSV filename
    project_id = xlsx_path.stem.split()[0] if " " in xlsx_path.stem else xlsx_path.stem
    csv_name = f"istha_{open_date}_{project_id}_{begin_date}.csv"
    csv_path = output_dir / csv_name

    rows_written = 0
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(ISTHA_HEADERS)

        for row_idx in range(7, ws.max_row + 1):
            code = ws.cell(row=row_idx, column=1).value
            first_cell = str(code).strip().lower() if code else ""
            quantity = ws.cell(row=row_idx, column=4).value

            # Skip header/empty/zero rows
            if not first_cell or first_cell.startswith(("item", "number")) or not quantity:
                continue

            abbreviation = str(ws.cell(row=row_idx, column=2).value or "").replace("  ", " ")
            unit = str(ws.cell(row=row_idx, column=3).value or "")

            for n, contractor_name in enumerate(contractors):
                price = ws.cell(row=row_idx, column=5 + n * 2).value or 0
                writer.writerow([
                    contractor_name, n + 1, code, abbreviation, unit,
                    quantity, price, bidder_totals[n], auth_totals[n], project_no,
                ])
                rows_written += 1

    wb.close()
    logger.info(f"Converted {xlsx_path.name} → {csv_name} ({rows_written} rows)")
    return csv_path


# ---------------------------------------------------------------------------
# Bid tab file renaming
# ---------------------------------------------------------------------------

def rename_bid_tab_file(content: str) -> str:
    """
    Generate a standardized filename from IDOT bid tab file content.

    Parses letting date and contract number from the fixed-width header.

    Input content: raw text of an IDOT bid tab file
    Output: e.g., "SCH20030117ILTABS44800.txt"
    """
    lines = content.split("\n")

    # Fix leading-space anomaly
    if len(lines) > 2 and lines[2].startswith(" "):
        lines = [line[1:] if line.startswith(" ") else line for line in lines]

    letting_date = lines[2][14:24].strip()
    letting_type = lines[2][40:43].strip()
    contract = lines[2][99:105].strip()

    parts = letting_date.split("/")
    return f"{letting_type}{parts[2]}{parts[0]}{parts[1]}ILTABS{contract}.txt"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Patterns for extracting dates from filenames
RE_DATE_YYYY_MM_DD = re.compile(r"(\d{4})_(\d{2})_(\d{2})")
RE_DATE_MMDDYY = re.compile(r"(\d{2})(\d{2})(\d{2})")
RE_DATE_LEGACY = re.compile(r"APIR(\d{6})")


def _detect_format(ws) -> list[int]:
    """
    Detect which IDOT Excel format a worksheet uses by inspecting headers.

    Format A/B (2018–2022): headers include "COUNTY" before "DIST"
      → county is col 4, district is col 5
    Format C (2023+): headers include "District" before "Contract", "County" last
      → district is col 4, county is col 9
    """
    # Collect first 3 rows for analysis
    header_rows = []
    for row in ws.iter_rows(max_row=3, values_only=True):
        if row:
            header_rows.append([str(c).strip().lower() if c else "" for c in row])

    # Search all header rows for format markers
    for cells in header_rows:
        # Format C: has "awarded" column (2023+ format)
        if "awarded" in cells:
            return _COL_MAP_FORMAT_C

    for cells in header_rows:
        # Format A/B: "county" appears in position 4
        if len(cells) > 5 and cells[4] == "county":
            return _COL_MAP_FORMAT_A

    # Default to Format A (older, more common)
    logger.warning("Could not detect Excel format, defaulting to Format A (9-col)")
    return _COL_MAP_FORMAT_A


def _parse_date_from_filename(stem: str) -> datetime | None:
    """
    Extract a date from various filename conventions:
      - IDOT_PayItemAwards_2025_11_07  → 2025-11-07
      - APIR110725                     → 2025-11-07  (legacy MMDDYY)
      - AWD_IL_IDOT_2025_11_07         → 2025-11-07
    """
    # Try YYYY_MM_DD pattern first
    match = RE_DATE_YYYY_MM_DD.search(stem)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            pass

    # Try legacy APIR + MMDDYY
    match = RE_DATE_LEGACY.search(stem)
    if match:
        try:
            return datetime.strptime(match.group(1), "%m%d%y")
        except ValueError:
            pass

    # Try bare MMDDYY at end of stem
    match = RE_DATE_MMDDYY.search(stem[-6:])
    if match:
        try:
            return datetime.strptime(match.group(0), "%m%d%y")
        except ValueError:
            pass

    return None


def _clean_cell(value) -> str:
    """Clean a cell value for CSV output."""
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="IDOT/ISTHA Excel → CSV converter")
    parser.add_argument("input_dir", help="Directory containing .xlsx files")
    parser.add_argument("--output-dir", default=None, help="Output directory (default: same as input)")
    parser.add_argument("--type", choices=["idot-awards", "istha"], default="idot-awards",
                        help="File type to convert")
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    if args.type == "idot-awards":
        results = batch_convert_idot_awards(args.input_dir, args.output_dir)
    else:
        input_dir = Path(args.input_dir)
        results = []
        for f in sorted(input_dir.glob("*.xlsx")):
            try:
                results.append(istha_xlsx_to_csv(f, args.output_dir or input_dir))
            except Exception as e:
                logger.error(f"Failed: {f.name}: {e}")

    print(f"\nConverted {len(results)} files")
